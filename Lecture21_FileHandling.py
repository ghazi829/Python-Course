# import xlrd
import openpyxl


loc = openpyxl.load_workbook("studentsdetails.xlsx")
mysheet = loc["record"]
myvalue = mysheet.cell(4,3).value
print(myvalue)
print("=================================================")
# for x in range(4,mysheet.max_row+1):
ali_total = 0
std_name = mysheet.cell(4,3).value
print(std_name,end=" ")
max = 0
for x in range(4,mysheet.max_column+1):
    # print(mysheet.cell(x,3).value," - ",mysheet.cell(x,4).value)
    print(mysheet.cell(4,x).value,end=" ")
    ali_total += mysheet.cell(4,x).value
    if mysheet.cell(4,x).value > max:
        max = mysheet.cell(4,x).value
print(ali_total,max)



